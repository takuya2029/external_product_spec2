"""対外製品規格作成アプリ

起動: streamlit run external_product_spec_app.py
必要: streamlit, openpyxl, pypdf
"""
from __future__ import annotations

import datetime as dt
import io
import re
from copy import copy
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
import streamlit as st
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from pypdf import PdfReader


FIELDS: tuple[tuple[str, str], ...] = (
    ("product_name", "開発製品名"),
    ("control_number", "整理番号"),
    ("revision_date", "作成日／改定日"),
    ("appearance", "外観"),
    ("ph", "液性（pH）"),
    ("heavy_metals", "重金属"),
    ("arsenic", "ヒ素"),
    ("methanol", "メタノール"),
    ("phosphate", "全リン酸塩"),
)
CUSTOM_FIELD_ROWS = 8

# 整理番号・作成日／改定日は内部の抽出値としては保持するが、利用者が確認・編集する項目には表示しない。
FORM_FIELDS = tuple((key, label) for key, label in FIELDS if key not in {"control_number", "revision_date"})

# 各社の帳票表記の違いを吸収するため、検索語は複数指定しています。
LABELS: dict[str, tuple[str, ...]] = {
    # 「製品名」「品名」は帳票一覧の列見出しにも存在するため使わない。
    "product_name": ("開発製品名", "開発品名"),
    "control_number": ("整理番号", "管理番号", "規格番号"),
    "revision_date": ("作成日", "改定日", "制定日", "作成・改定日"),
    "appearance": ("外観",),
    "ph": ("液性", "pH", "ｐＨ"),
    "heavy_metals": ("重金属",),
    "arsenic": ("ヒ素", "砒素"),
    "methanol": ("メタノール", "Methanol"),
    "phosphate": ("リン酸塩", "りん酸塩", "全リン"),
}


def normalize(value: Any) -> str:
    """比較用に表記ゆれ、空白、全角英数をならす。"""
    text = "" if value is None else str(value)
    text = text.replace("\u3000", " ").replace("\n", " ")
    text = text.translate(str.maketrans("ｐＰｈＨ％", "pPhH%"))
    return re.sub(r"\s+", "", text).lower()


def display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.date, dt.datetime)):
        return value.strftime("%Y/%m/%d")
    return str(value).strip()


def is_label(value: Any, aliases: tuple[str, ...]) -> bool:
    cell = normalize(value)
    return bool(cell) and any(normalize(alias) in cell for alias in aliases)


def candidates_near_label(ws, row: int, col: int) -> list[str]:
    """同じ行のラベル右側だけを、近い順に返す。

    規格表は項目が縦に並ぶため、下方向を探索すると次の項目
    （例: 次の「冷時安定性」）を誤って採用してしまう。
    """
    candidates: list[str] = []
    for offset in range(1, 7):
        cell = ws.cell(row, col + offset)
        value = display_value(cell.value)
        if value and not is_label(value, tuple(a for xs in LABELS.values() for a in xs)):
            candidates.append(value)
    return candidates


def is_document_code(value: str) -> bool:
    """帳票番号・規格番号は品質規格の値ではないため除外する。"""
    compact = re.sub(r"\s+", "", value).upper()
    return bool(re.fullmatch(r"(?:[A-Z]{2,}\d+[A-Z\d-]*|\d{5,}-?(?:VER\.?\d+)?)", compact))


def is_plausible_value(key: str, value: str) -> bool:
    """帳票一覧や目次の文字列を規格値として採用しないための項目別検証。"""
    value = value.strip()
    if not value or len(value) > 120:
        return False
    if key == "revision_date":
        return bool(re.search(r"(?:19|20)?\d{2}[/.年-]\d{1,2}[/.月-]\d{1,2}|\d{6}[- ]?Ver\.?\d+", value, re.I))
    if is_document_code(value):
        return False
    if key == "ph":
        numbers = [float(n.replace("．", ".")) for n in re.findall(r"\d+(?:[.．]\d+)?", value)]
        return bool(numbers) and all(0 < n <= 14 for n in numbers) and not bool(re.search(r"NSPF|JIS", value, re.I))
    if key == "appearance":
        return bool(re.search(r"透明|液体|液状|固体|粉末|白色|無色|淡[黄褐色]|着色|ペースト", value))
    if key in {"heavy_metals", "arsenic", "methanol", "phosphate"}:
        return bool(re.search(r"(?:以下|未満|不検出|陰性|なし|せず|適合|準じる|mg|ppm|%|\d)", value, re.I))
    return True


def extract_ph_condition(label: str) -> str:
    """製品規格のpH項目名から、濃度・温度などの測定条件を取り出す。"""
    text = str(label).replace("\n", " ").strip()
    match = re.search(r"(?:pH|ｐＨ)\s*[（(]\s*([^）)]+)\s*[）)]", text, re.I)
    return match.group(1).strip() if match else ""


PH_RANGE_PATTERN = re.compile(
    r"(?P<low>\d+(?:[.．]\d+)?)\s*(?:～|〜|~|－|−|–|—|-)\s*(?P<high>\d+(?:[.．]\d+)?)"
)


def format_decimal(value: Decimal) -> str:
    """不要な末尾のゼロを除いた10進表記にする。"""
    text = format(value.normalize(), "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def format_ph_range(value: str) -> str:
    """液性の範囲を「中心値±幅」に変換する。例: 10.5～11.1 → 10.8±0.3。"""
    def replace(match: re.Match[str]) -> str:
        try:
            low = Decimal(match.group("low").replace("．", "."))
            high = Decimal(match.group("high").replace("．", "."))
        except InvalidOperation:
            return match.group(0)
        center = (low + high) / Decimal("2")
        width = abs(high - low) / Decimal("2")
        return f"{format_decimal(center)}±{format_decimal(width)}"

    return PH_RANGE_PATTERN.sub(replace, value, count=1)


PPM_OR_MG_PER_KG_PATTERN = re.compile(
    r"(?P<number>\d+(?:[.．]\d+)?)\s*(?:ppm|ｐｐｍ|ＰＰＭ|mg\s*/\s*kg|ｍｇ\s*/\s*ｋｇ)",
    re.I,
)
BARE_NUMBER_PATTERN = re.compile(r"(?P<number>\d+(?:[.．]\d+)?)")
OTHER_UNIT_PATTERN = re.compile(r"(?:mg\s*/\s*(?!kg)|ｍｇ\s*/\s*(?!ｋｇ)|g\s*/|ｇ\s*/|μg|ug|％|%)", re.I)


def format_ppm_value(value: str) -> str:
    """ヒ素・重金属の濃度をppm表記へそろえ、数値との間を半角スペースにする。"""
    value = PPM_OR_MG_PER_KG_PATTERN.sub(r"\g<number> ppm", value)
    if re.search(r"ppm", value, re.I):
        return value
    # ppmと等価ではない単位（mg/L、%、μgなど）は、元の製品規格の表記を保持する。
    if OTHER_UNIT_PATTERN.search(value):
        return value
    # 単位が省略されている数値にはppmを補う。不検出など数値のない値は変更しない。
    return BARE_NUMBER_PATTERN.sub(r"\g<number> ppm", value, count=1)


def decimal_from_text(value: str) -> Decimal:
    return Decimal(value.replace("．", "."))


def methanol_limit_mg_per_g(value: str) -> Decimal | None:
    """「以下」のメタノール規格値をmg/gへ換算する。換算不能ならNoneを返す。"""
    if "以下" not in value:
        return None
    number = r"(?P<value>\d+(?:[.．]\d+)?)"
    match = re.search(rf"{number}\s*(?:mg|ｍｇ)\s*/\s*(?:g|ｇ)", value, re.I)
    if match:
        return decimal_from_text(match.group("value"))
    match = re.search(
        r"(?P<grams>\d+(?:[.．]\d+)?)\s*(?:g|ｇ)\s*中\s*"
        r"(?P<milligrams>\d+(?:[.．]\d+)?)\s*(?:mg|ｍｇ)",
        value,
        re.I,
    )
    if match:
        return decimal_from_text(match.group("milligrams")) / decimal_from_text(match.group("grams"))
    match = re.search(rf"{number}\s*(?:ppm|ｐｐｍ|ＰＰＭ)", value, re.I)
    if match:
        return decimal_from_text(match.group("value")) / Decimal("1000")
    match = re.search(rf"{number}\s*(?:mg|ｍｇ)\s*/\s*(?:kg|ｋｇ)", value, re.I)
    if match:
        return decimal_from_text(match.group("value")) / Decimal("1000")
    match = re.search(rf"{number}\s*(?:%|％)", value)
    if match:
        return decimal_from_text(match.group("value")) * Decimal("10")
    return None


def format_methanol_value(value: str) -> str:
    """1 mg/g以下と同値の抽出値だけを、対外規格用の表記へ統一する。"""
    limit = methanol_limit_mg_per_g(value)
    return "1 mg/g以下" if limit == Decimal("1") else value


def extract_from_workbook(raw: bytes, filename: str) -> dict[str, str]:
    keep_vba = filename.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, keep_vba=keep_vba)
    found = {key: "" for key, _ in FIELDS}
    # 帳票一覧・改訂履歴からの抽出は誤認の原因になるので、規格シートだけを対象にする。
    spec_sheets = [ws for ws in wb.worksheets if "製品規格" in str(ws.title)]
    sheets = spec_sheets or [ws for ws in wb.worksheets if "規格" in str(ws.title)] or [wb.active]
    for ws in sheets:
        # 製品規格の標準帳票: B列「品質特性」、E列「製品規格」の表を優先する。
        # 試験方法列（NSPF...）ではなく、必ず「製品規格」列を読取る。
        quality_header = next(
            (cell for row in ws.iter_rows() for cell in row if "品質特性" in normalize(cell.value)), None
        )
        # シート題名の「製品規格」（A3）ではなく、品質表の同じ見出し行を選ぶ。
        spec_header = next(
            (cell for cell in ws[quality_header.row] if normalize(cell.value) == normalize("製品規格")), None
        ) if quality_header else None
        if quality_header and spec_header:
            table_start = max(quality_header.row, spec_header.row) + 1
            rows = [(display_value(ws.cell(r, quality_header.column).value),
                     display_value(ws.cell(r, spec_header.column).value))
                    for r in range(table_start, min(ws.max_row, table_start + 80) + 1)]
            row_rules = {
                "appearance": lambda label: "外観" in normalize(label),
                "heavy_metals": lambda label: "重金属" in normalize(label),
                "arsenic": lambda label: "ひ素" in normalize(label) or "砒素" in normalize(label),
                "methanol": lambda label: "メタノール" in normalize(label),
                "phosphate": lambda label: "リン酸塩" in normalize(label),
            }
            for key, matches in row_rules.items():
                if not found[key]:
                    found[key] = next((value for label, value in rows if matches(label) and value), "")

            # pHは濃度を固定せず、規格表内で有効な値を持つpH行を採用する。
            if not found["ph"]:
                ph_row = next(
                    ((label, value) for label, value in rows
                     if "ph" in normalize(label) and is_plausible_value("ph", value)),
                    None,
                )
                if ph_row:
                    found["ph"] = ph_row[1]
                    found["ph_condition"] = extract_ph_condition(ph_row[0])

            # 帳票ヘッダーから製品名、規格番号、改定番号を取得する。
            header_values = [display_value(cell.value) for row in ws.iter_rows(min_row=1, max_row=12) for cell in row]
            title_cell = next(
                (cell for row in ws.iter_rows(min_row=1, max_row=12) for cell in row
                 if normalize(cell.value) == normalize("製品規格")), None
            )
            if title_cell:
                found["product_name"] = display_value(
                    ws.cell(title_cell.row + 1, title_cell.column + 1).value
                ) or found["product_name"]
            if not found["product_name"]:
                # 「○○○○ 製品規格」のように、帳票タイトルへ製品名が含まれる形式を補完する。
                title_with_name = next(
                    (
                        display_value(cell.value)
                        for row in ws.iter_rows(min_row=1, max_row=12)
                        for cell in row
                        if "製品規格" in display_value(cell.value)
                        and normalize(cell.value) != normalize("製品規格")
                    ),
                    "",
                )
                found["product_name"] = re.sub(r"\s*製品規格\s*$", "", title_with_name).strip()
            found["control_number"] = next(
                (value for value in header_values if re.fullmatch(r"[ＮN][ＳS][ＰP][ＫK][０-９0-9Ａ-ＺA-Z]+", value)),
                found["control_number"],
            )
            found["revision_date"] = next(
                (value for value in header_values if re.fullmatch(r"\d{6}[-－]Ver\.?[\d.]+", value, re.I)),
                found["revision_date"],
            )
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                for key, _ in FIELDS:
                    if not found[key] and is_label(cell.value, LABELS[key]):
                        for candidate in candidates_near_label(ws, cell.row, cell.column):
                            if is_plausible_value(key, candidate):
                                found[key] = candidate
                                break
    return found


def extract_pdf_text(raw: bytes) -> str:
    return "\n".join((page.extract_text() or "") for page in PdfReader(io.BytesIO(raw)).pages)


def extract_from_pdf(raw: bytes) -> dict[str, str]:
    text = extract_pdf_text(raw)
    result = {key: "" for key, _ in FIELDS}
    for key, _ in FIELDS:
        for alias in LABELS[key]:
            # 「ラベル : 値」とラベルの直後の行の両方を受け付ける。
            match = re.search(rf"{re.escape(alias)}\s*(?:[:：]|\t|\s{{2,}})\s*([^\n]+)", text, re.I)
            if not match:
                match = re.search(rf"{re.escape(alias)}\s*\n\s*([^\n]+)", text, re.I)
            if match:
                value = match.group(1).strip(" ：:\t")
                if value and len(value) < 160:
                    result[key] = value
                    break
    return result


def extract_spec(raw: bytes, filename: str) -> dict[str, str]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return extract_from_workbook(raw, filename)
    if suffix == ".pdf":
        return extract_from_pdf(raw)
    raise ValueError("製品規格は .xlsx / .xlsm / .pdf を指定してください。")


def writable_cell(ws, row: int, col: int):
    cell = ws.cell(row, col)
    return None if isinstance(cell, MergedCell) else cell


def put_near_label(ws, aliases: tuple[str, ...], value: str) -> bool:
    if not value:
        return False
    for row in ws.iter_rows():
        for label_cell in row:
            if not is_label(label_cell.value, aliases):
                continue
            # フォーマットでは通常、ラベルの右が入力欄。結合セルも安全に避ける。
            for offset in range(1, 9):
                target = writable_cell(ws, label_cell.row, label_cell.column + offset)
                if target is not None:
                    target.value = value
                    return True
            for offset in range(1, 4):
                target = writable_cell(ws, label_cell.row + offset, label_cell.column)
                if target is not None:
                    target.value = value
                    return True
    return False


def replace_placeholders(ws, values: dict[str, str]) -> set[str]:
    """{{product_name}} のような明示プレースホルダーもサポートする。"""
    written: set[str] = set()
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            for key, value in values.items():
                token = "{{" + key + "}}"
                if token in cell.value and value:
                    cell.value = cell.value.replace(token, value)
                    written.add(key)
    return written


def write_custom_fields(ws, custom_fields: list[tuple[str, str]]) -> list[tuple[str, str]]:
    """任意項目の {{項目名}} プレースホルダーを置換し、未配置項目を返す。"""
    unwritten: list[tuple[str, str]] = []
    for label, value in custom_fields:
        label, value = label.strip(), value.strip()
        if not label or not value:
            continue
        placeholder_written = False
        token = "{{" + label + "}}"
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and token in cell.value:
                    cell.value = cell.value.replace(token, value)
                    placeholder_written = True
        if not placeholder_written:
            unwritten.append((label, value))
    return unwritten


def append_custom_rows(ws, custom_fields: list[tuple[str, str]]) -> None:
    """ラベルがない任意項目を、品質特性表の末尾（備考の前）へ追加する。"""
    if not custom_fields:
        return
    quality_header = next(
        (cell for row in ws.iter_rows() for cell in row if "品質特性" in normalize(cell.value)), None
    )
    if quality_header is None:
        return
    note_row = next(
        (
            row for row in range(quality_header.row + 1, ws.max_row + 1)
            if "備考" in normalize(ws.cell(row, quality_header.column).value)
        ),
        None,
    )
    if note_row is None or note_row <= quality_header.row + 1:
        return

    # 既存の結合範囲を一時的に解除してから行を挿入し、備考以下の結合範囲を正しい行へ戻す。
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))
    ws.insert_rows(note_row, amount=len(custom_fields))
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )
        if min_row >= note_row:
            min_row += len(custom_fields)
            max_row += len(custom_fields)
        ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

    style_row = note_row - 1
    for offset, (label, value) in enumerate(custom_fields):
        target_row = note_row + offset
        for column in range(1, ws.max_column + 1):
            source = ws.cell(style_row, column)
            target = ws.cell(target_row, column)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format
            target.alignment = copy(source.alignment)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.protection = copy(source.protection)
        ws.row_dimensions[target_row].height = ws.row_dimensions[style_row].height
        ws.merge_cells(start_row=target_row, start_column=1, end_row=target_row, end_column=2)
        ws.merge_cells(start_row=target_row, start_column=3, end_row=target_row, end_column=5)
        # Excelは境界線で前行の下罫線を優先表示するため、前行の下罫線と
        # 追加行の上罫線の両方を周囲と同じ点線にする。
        for column in range(1, 6):
            previous_border = copy(ws.cell(target_row - 1, column).border)
            previous_border.bottom = Side(style="dotted", color="000000")
            ws.cell(target_row - 1, column).border = previous_border
            border = copy(ws.cell(target_row, column).border)
            border.top = Side(style="dotted", color="000000")
            ws.cell(target_row, column).border = border
        ws.cell(target_row, 1).value = label
        ws.cell(target_row, 3).value = value

    # 追加項目の最終行と備考の境界は、表の区切りとして実線を維持する。
    last_custom_row = note_row + len(custom_fields) - 1
    note_after_row = note_row + len(custom_fields)
    for column in range(1, 6):
        last_border = copy(ws.cell(last_custom_row, column).border)
        last_border.bottom = Side(style="thin", color="000000")
        ws.cell(last_custom_row, column).border = last_border
        note_border = copy(ws.cell(note_after_row, column).border)
        note_border.top = Side(style="thin", color="000000")
        ws.cell(note_after_row, column).border = note_border


def find_product_list_sheet(wb):
    return next((ws for ws in wb.worksheets if "製品一覧" in normalize(ws.title)), None)


def product_list_header(ws) -> tuple[int, dict[str, int]]:
    """製品一覧の見出し行と、見出し名→列番号を返す。"""
    for row_index in range(1, min(ws.max_row, 10) + 1):
        columns = {
            normalize(cell.value): cell.column
            for cell in ws[row_index]
            if display_value(cell.value)
        }
        if "製品名" in columns:
            return row_index, columns
    return 0, {}


def is_placeholder_product(value: str) -> bool:
    return not value or "製品名を記入" in value or "○○○" in value


def clean_product_name(value: str) -> str:
    """製品一覧では表題用の「製品規格」という接尾語を保持しない。"""
    return re.sub(r"\s*製品規格\s*$", "", value).strip()


def is_non_product_value(value: str) -> bool:
    """製品一覧の末尾にある日付・備考行を製品として扱わない。"""
    return value in {"日付", "備考"} or bool(re.fullmatch(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}(?:\s+\d{1,2}:\d{2}:\d{2})?", value))


def product_key(value: str) -> str:
    """「リキッドPH」と「リキッドPH 製品規格」を同一製品として扱う。"""
    return normalize(clean_product_name(value))


def product_row_score(ws, row: int, columns: dict[str, int]) -> int:
    """ダミー値を含まない行を優先して残すための品質スコア。"""
    values = [display_value(ws.cell(row, col).value) for col in columns.values()]
    return sum(1 for value in values if value and "○" not in value and "製品名を記入" not in value)


def deduplicate_product_list(ws, header_row: int, columns: dict[str, int]) -> None:
    product_col = columns["製品名"]
    retained: dict[str, int] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        product_name = clean_product_name(display_value(ws.cell(row, product_col).value))
        ws.cell(row, product_col).value = product_name
        key = product_key(product_name)
        if is_non_product_value(product_name):
            # 製品表の途中へ誤転記された日付値は削除する（本来の「日付」見出し行は保持）。
            if re.fullmatch(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}(?:\s+\d{1,2}:\d{2}:\d{2})?", product_name):
                for col in columns.values():
                    ws.cell(row, col).value = ""
            continue
        if not key or is_placeholder_product(product_name):
            continue
        if key not in retained:
            retained[key] = row
            continue
        old_row = retained[key]
        keep_row, clear_row = (row, old_row) if product_row_score(ws, row, columns) > product_row_score(ws, old_row, columns) else (old_row, row)
        for col in columns.values():
            ws.cell(clear_row, col).value = ""
        retained[key] = keep_row


def merge_product_lists(output_wb, past_files: list[tuple[bytes, str]]) -> int:
    """過去帳票の製品一覧を、製品名の重複を避けて出力帳票へ転記する。"""
    target_ws = find_product_list_sheet(output_wb)
    if target_ws is None:
        return 0
    target_header_row, target_columns = product_list_header(target_ws)
    product_column = target_columns.get("製品名")
    if not target_header_row or product_column is None:
        return 0

    existing_names = {
        product_key(display_value(target_ws.cell(row, product_column).value))
        for row in range(target_header_row + 1, target_ws.max_row + 1)
        if not is_placeholder_product(display_value(target_ws.cell(row, product_column).value))
    }
    copied = 0
    for raw, filename in past_files:
        try:
            source_wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, keep_vba=filename.lower().endswith(".xlsm"))
        except Exception:
            continue
        source_ws = find_product_list_sheet(source_wb)
        if source_ws is None:
            continue
        source_header_row, source_columns = product_list_header(source_ws)
        source_product_col = source_columns.get("製品名")
        if not source_header_row or source_product_col is None:
            continue
        for source_row in range(source_header_row + 1, source_ws.max_row + 1):
            product_name = clean_product_name(display_value(source_ws.cell(source_row, source_product_col).value))
            if is_placeholder_product(product_name) or is_non_product_value(product_name) or product_key(product_name) in existing_names:
                continue
            # 空欄またはテンプレートのダミー行を再利用して、製品一覧を詰めて作成する。
            target_row = next(
                (row for row in range(target_header_row + 1, target_ws.max_row + 2)
                 if is_placeholder_product(display_value(target_ws.cell(row, product_column).value))),
                target_ws.max_row + 1,
            )
            for header, target_col in target_columns.items():
                source_col = source_columns.get(header)
                if source_col:
                    value = source_ws.cell(source_row, source_col).value
                    target_ws.cell(target_row, target_col).value = product_name if header == normalize("製品名") else value
            existing_names.add(product_key(product_name))
            copied += 1
    deduplicate_product_list(target_ws, target_header_row, target_columns)
    # VLOOKUPの参照範囲を、転記後の製品一覧の末尾まで拡張する。
    last_row = max(target_header_row + 1, target_ws.max_row)
    for ws in output_wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "VLOOKUP(" in cell.value.upper() and "製品一覧!" in cell.value:
                    cell.value = re.sub(r"製品一覧!\$?A\$?2:\$?F\$?\d+", f"製品一覧!A2:F{last_row}", cell.value)
    return copied


def set_product_selector(wb) -> None:
    """対外製品規格シートの製品名・日付セルへリスト入力規則を設定する。"""
    product_list_ws = find_product_list_sheet(wb)
    target_ws = next((ws for ws in wb.worksheets if "対外製品規格" in normalize(ws.title)), None)
    if product_list_ws is None or target_ws is None:
        return
    # 他シート参照のリストはExcelで安定して使えるよう、名前付き範囲を介して指定する。
    range_name = "external_product_names"
    if range_name in wb.defined_names:
        del wb.defined_names[range_name]
    wb.defined_names.add(DefinedName(range_name, attr_text="'製品一覧'!$A$2:$A$17"))
    validation = DataValidation(type="list", formula1=f"={range_name}", allow_blank=True)
    validation.error = "製品一覧のリストから製品名を選択してください。"
    validation.errorTitle = "製品名の入力"
    target_ws.add_data_validation(validation)
    validation.add(target_ws["C1"])

    date_range_name = "external_spec_dates"
    if date_range_name in wb.defined_names:
        del wb.defined_names[date_range_name]
    wb.defined_names.add(DefinedName(date_range_name, attr_text="'製品一覧'!$A$23:$A$24"))
    date_validation = DataValidation(type="list", formula1=f"={date_range_name}", allow_blank=True)
    date_validation.error = "製品一覧の日付リストから選択してください。"
    date_validation.errorTitle = "日付の入力"
    target_ws.add_data_validation(date_validation)
    date_validation.add(target_ws["E3"])
    # 出力時点の日付を入れておくため、プルダウンを選択し直す必要をなくす。
    target_ws["E3"] = dt.date.today()


def set_company_details(wb, product_name: str) -> None:
    """製品一覧の会社情報を、出力帳票へ直接書き込む。"""
    product_list_ws = find_product_list_sheet(wb)
    target_ws = next((ws for ws in wb.worksheets if "対外製品規格" in normalize(ws.title)), None)
    if product_list_ws is None or target_ws is None or not product_name:
        return
    header_row, columns = product_list_header(product_list_ws)
    product_column = columns.get(normalize("製品名"))
    if not header_row or product_column is None:
        return
    source_row = next(
        (
            row for row in range(header_row + 1, product_list_ws.max_row + 1)
            if product_key(display_value(product_list_ws.cell(row, product_column).value)) == product_key(product_name)
        ),
        None,
    )
    if source_row is None:
        return
    for cell_ref, header in (("E5", "会社名"), ("E6", "住所"), ("E7", "電話番号")):
        source_column = columns.get(normalize(header))
        if source_column is not None:
            target_ws[cell_ref] = product_list_ws.cell(source_row, source_column).value


def select_first_product(wb, preferred_product_name: str = "") -> str:
    """抽出製品名を優先し、なければ製品一覧の先頭名を対外規格の選択値へ設定する。"""
    product_list_ws = find_product_list_sheet(wb)
    target_ws = next((ws for ws in wb.worksheets if "対外製品規格" in normalize(ws.title)), None)
    if product_list_ws is None or target_ws is None:
        return ""
    header_row, columns = product_list_header(product_list_ws)
    product_column = columns.get(normalize("製品名"))
    if not header_row or product_column is None:
        return ""
    product_name = clean_product_name(preferred_product_name)
    if not product_name:
        product_name = next(
            (
                clean_product_name(display_value(product_list_ws.cell(row, product_column).value))
                for row in range(header_row + 1, product_list_ws.max_row + 1)
                if not is_placeholder_product(display_value(product_list_ws.cell(row, product_column).value))
                and not is_non_product_value(display_value(product_list_ws.cell(row, product_column).value))
            ),
            "",
        )
    if product_name:
        target_ws["C1"] = product_name
    return product_name


def align_transferred_values(wb) -> None:
    """対外規格の自動転記欄を、全角スペース1文字で字下げしてそろえる。"""
    target_ws = next((ws for ws in wb.worksheets if "対外製品規格" in normalize(ws.title)), None)
    if target_ws is None:
        return
    # 品質特性ラベル（A10:A23）と規格値（C10:C23）を同じ字下げ幅にする。
    for cell_ref in [*[f"A{row}" for row in range(10, 24)], *[f"C{row}" for row in range(10, 24)]]:
        cell = target_ws[cell_ref]
        if not isinstance(cell.value, str) or cell.value.startswith("="):
            continue
        # 複数行セルも、全ての行をそろえる。
        cell.value = "\n".join("　" + line.lstrip("　 ") for line in cell.value.splitlines())


def fill_template(
    raw: bytes,
    filename: str,
    values: dict[str, str],
    past_files: list[tuple[bytes, str]] | None = None,
    custom_fields: list[tuple[str, str]] | None = None,
) -> bytes:
    keep_vba = filename.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(io.BytesIO(raw), keep_vba=keep_vba)
    values = {
        **values,
        "ph": format_ph_range(values.get("ph", "")),
        "heavy_metals": format_ppm_value(values.get("heavy_metals", "")),
        "arsenic": format_ppm_value(values.get("arsenic", "")),
        "methanol": format_methanol_value(values.get("methanol", "")),
    }
    done: set[str] = set()
    for ws in wb.worksheets:
        done |= replace_placeholders(ws, values)
        for key, _ in FIELDS:
            if key not in done and put_near_label(ws, LABELS[key], values.get(key, "")):
                done.add(key)
        unwritten_custom_fields = write_custom_fields(ws, custom_fields or [])
        if "対外製品規格" in normalize(ws.title):
            append_custom_rows(ws, unwritten_custom_fields)
        # 「○○○○ 製品規格」のような表題へ、開発製品名を確実に反映する。
        product_name = values.get("product_name", "")
        if product_name and "製品一覧" not in normalize(ws.title):
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)):
                for cell in row:
                    if isinstance(cell.value, str) and "製品規格" in cell.value:
                        # C1は製品一覧のドロップダウンと同じ製品名を保持する。
                        cell.value = product_name
                        break
                else:
                    continue
                break
        # 製品規格のpH項目名から抽出した濃度・温度条件を、対外規格の項目名へ反映する。
        for row in ws.iter_rows():
            for cell in row:
                if (
                    isinstance(cell.value, str)
                    and "液性" in normalize(cell.value)
                    and "ph" in normalize(cell.value)
                ):
                    condition = values.get("ph_condition") or "0.20％水溶液、25℃"
                    cell.value = f"液性(pH、{condition})"
    if past_files:
        merge_product_lists(wb, past_files)
    set_product_selector(wb)
    selected_product_name = select_first_product(wb, values.get("product_name", ""))
    set_company_details(wb, selected_product_name)
    align_transferred_values(wb)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def init_form() -> None:
    for key, _ in FORM_FIELDS:
        st.session_state.setdefault(f"form_{key}", "")
    st.session_state.setdefault("form_ph_condition", "")
    for row in range(1, CUSTOM_FIELD_ROWS + 1):
        st.session_state.setdefault(f"custom_label_{row}", "")
        st.session_state.setdefault(f"custom_value_{row}", "")


def apply_extracted(data: dict[str, str]) -> None:
    for key, value in data.items():
        if key == "ph_condition" and value:
            st.session_state["form_ph_condition"] = value
        if key == "ph":
            value = format_ph_range(value)
        if key in {"heavy_metals", "arsenic"}:
            value = format_ppm_value(value)
        if value and any(key == form_key for form_key, _ in FORM_FIELDS):
            st.session_state[f"form_{key}"] = value


st.set_page_config(page_title="対外製品規格作成", page_icon="📄", layout="wide")
st.markdown(
    """
    <style>
        .block-container {max-width: 1180px; padding-top: 2.4rem; padding-bottom: 3rem;}
        #MainMenu, footer {visibility: hidden;}
        .hero {padding: 2rem 2.2rem; border-radius: 20px; color: #fff;
               background: linear-gradient(120deg, #0f3d5e 0%, #176b87 56%, #38a3a5 100%);
               box-shadow: 0 14px 30px rgba(15, 61, 94, 0.18); margin-bottom: 1.4rem;}
        .hero h1 {margin: 0 0 .45rem; font-size: 2rem; letter-spacing: .04em;}
        .hero p {margin: 0; opacity: .9; font-size: 1rem;}
        .step {margin: 1.7rem 0 .7rem; color: #176b87; font-size: .82rem; font-weight: 700;
               letter-spacing: .12em;}
        .step span {color: #1f2937; font-size: 1.22rem; letter-spacing: 0; margin-left: .55rem;}
        .guide {background: #f0f8f9; border-left: 4px solid #38a3a5; border-radius: 8px;
                padding: .75rem 1rem; color: #3d5561; margin-bottom: .8rem;}
        [data-testid="stFileUploader"] {padding: .4rem .2rem;}
        [data-testid="stDownloadButton"] button, [data-testid="stButton"] button {
            border-radius: 9px; font-weight: 650; min-height: 2.75rem;
        }
        [data-testid="stDownloadButton"] button {background: #176b87; color: white; border: 0;}
        [data-testid="stDownloadButton"] button:hover {background: #0f526b; color: white;}
        [data-testid="stTextInput"] input {border-radius: 8px;}
    </style>
    <div class="hero">
        <h1>対外製品規格作成</h1>
        <p>製品規格を読み込み、確認した内容を対外製品規格フォーマットへ転記します。</p>
    </div>
    """,
    unsafe_allow_html=True,
)
init_form()

with st.sidebar:
    st.markdown("### ご利用の流れ")
    st.markdown("**1.** 製品規格をアップロード  ")
    st.markdown("**2.** 抽出内容を確認・編集  ")
    st.markdown("**3.** フォーマットを指定してダウンロード")
    st.divider()
    st.caption("対応形式：Excel（.xlsx / .xlsm）・PDF")

st.markdown('<div class="step">STEP 01 <span>製品規格を読み込む</span></div>', unsafe_allow_html=True)
st.markdown('<div class="guide">製品規格ファイルをアップロード後、「抽出する」を押してください。</div>', unsafe_allow_html=True)
with st.expander("製品規格ファイルを選択", expanded=True):
    upload_col, action_col = st.columns([3, 1])
    with upload_col:
        spec = st.file_uploader("製品規格ファイル", type=["xlsx", "xlsm", "pdf"], label_visibility="collapsed")
    with action_col:
        extract_clicked = st.button("抽出する", type="primary", use_container_width=True)
    if extract_clicked:
        if spec is None:
            st.error("製品規格ファイルをアップロードしてください。")
        else:
            try:
                data = extract_spec(spec.getvalue(), spec.name)
                apply_extracted(data)
                st.success("抽出結果を確認フォームへ転記しました。空欄は手入力してください。")
            except Exception as exc:
                st.error(f"抽出できませんでした: {exc}")

st.markdown('<div class="step">STEP 02 <span>抽出内容を確認・編集</span></div>', unsafe_allow_html=True)
st.caption("空欄や表記を必要に応じて修正してください。")
left, right = st.columns(2)
for index, (key, label) in enumerate(FORM_FIELDS):
    container = left if index % 2 == 0 else right
    with container:
        if key == "ph":
            st.text_input("液性の測定条件", key="form_ph_condition", placeholder="例: 0.2%水溶液、25℃")
            condition = st.session_state.get("form_ph_condition", "")
            st.text_input(f"液性（pH、{condition or '測定条件未抽出'}）", key=f"form_{key}")
        else:
            st.text_input(label, key=f"form_{key}")

with st.expander("任意項目を追加（例外対応）"):
    st.caption("必要な行だけ入力してください。入力済みの行は項目に追加されます。")
    header_name, header_value = st.columns([2, 3])
    header_name.markdown("**項目名**")
    header_value.markdown("**規格**")
    for row in range(1, CUSTOM_FIELD_ROWS + 1):
        name_col, value_col = st.columns([2, 3])
        with name_col:
            st.text_input("項目名", key=f"custom_label_{row}", label_visibility="collapsed", placeholder="例: 蛍光増白剤")
        with value_col:
            st.text_input("規格値・内容", key=f"custom_value_{row}", label_visibility="collapsed", placeholder="例: 含有せず")

st.markdown('<div class="step">STEP 03 <span>対外製品規格を出力</span></div>', unsafe_allow_html=True)
st.markdown('<div class="guide">対外製品規格フォーマット、過去の対外製品規格をアップすると、対外製品規格フォーマットに自動転記します。</div>', unsafe_allow_html=True)
template_col, history_col = st.columns(2)
with template_col:
    template = st.file_uploader("対外製品規格フォーマット", type=["xlsx", "xlsm"])
with history_col:
    past_spec_files = st.file_uploader(
        "過去の対外製品規格",
        type=["xlsx", "xlsm"],
        accept_multiple_files=True,
        help="製品一覧へ追加したい過去の対外製品規格を選択します。",
    )
if template is not None:
    values = {key: st.session_state.get(f"form_{key}", "") for key, _ in FIELDS}
    values["ph_condition"] = st.session_state.get("form_ph_condition", "")
    custom_fields = [
        (st.session_state[f"custom_label_{row}"], st.session_state[f"custom_value_{row}"])
        for row in range(1, CUSTOM_FIELD_ROWS + 1)
        if st.session_state[f"custom_label_{row}"].strip() and st.session_state[f"custom_value_{row}"].strip()
    ]
    try:
        past_files = [(past_file.getvalue(), past_file.name) for past_file in past_spec_files]
        generated = fill_template(template.getvalue(), template.name, values, past_files, custom_fields)
        base = re.sub(r'[\\/:*?"<>|]', "_", values["product_name"] or "対外製品規格")
        suffix = Path(template.name).suffix.lower()
        st.download_button(
            "入力済み対外製品規格をダウンロード",
            data=generated,
            file_name=f"{base}_対外製品規格{suffix}",
            mime="application/vnd.ms-excel.sheet.macroEnabled.12" if suffix == ".xlsm" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    except Exception as exc:
        st.error(f"フォーマットへの転記に失敗しました: {exc}")
